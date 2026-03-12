import psycopg2
import psycopg2.extras
import json
import re
import os
import random
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

db_params = {
    "host":     "54.144.89.225",
    "database": "ai_butler_test",
    "user":     "app_user",
    "password": "App$gr0c",
    "port":     5432
}

C_DARK_BLUE  = "1F4E79"
C_MID_BLUE   = "2E75B6"
C_PALE_BLUE  = "EBF3FB"
C_PURPLE     = "4A148C"
C_LIGHT_PURP = "F3EBF9"
C_GREEN      = "1B5E20"
C_ORANGE     = "E65100"
C_LIGHT_ORG  = "FFF8F0"
C_WHITE      = "FFFFFF"

def fill(h):  return PatternFill("solid", fgColor=h)
def fnt(bold=False, color="000000", size=10, italic=False):
    return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)
def aln(h="left", wrap=True):
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)
def brdr(color="CCCCCC", thick=False):
    s = Side(style="medium" if thick else "thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def set_hdr(cell, text, bg, fg=C_WHITE):
    cell.value = text; cell.fill = fill(bg)
    cell.font = fnt(bold=True, color=fg, size=10)
    cell.alignment = aln("center", wrap=True)
    cell.border = brdr(bg, thick=True)

def set_cel(cell, value, bg=C_WHITE, bold=False, fg="222222", center=False, italic=False):
    cell.value = value; cell.fill = fill(bg)
    cell.font = fnt(bold=bold, color=fg, size=10, italic=italic)
    cell.alignment = aln("center" if center else "left", wrap=True)
    cell.border = brdr()

def widths(ws, d):
    for col, w in d.items():
        ws.column_dimensions[col].width = w

def parse_mins(t):
    try:    return int(re.sub(r"[^0-9]", "", str(t))) or None
    except: return None

def parse_amt(a):
    qty  = re.sub(r"[^0-9.]", "", str(a))
    unit = re.sub(r"[0-9.\s]", "", str(a))
    return (qty or str(a)), (unit or "")

def _safe_float(val):
    try:    return float(val) if val else 0.0
    except: return 0.0

def complexity(m):
    if m is None: return "Unknown"
    return "Easy" if m <= 20 else ("Medium" if m <= 40 else "Hard")

# ── DB fetch ─────────────────────────────────────────────────
def fetch_meal_plans():
    conn = psycopg2.connect(**db_params)
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    cur.execute("""
        SELECT id, user_id, prefs, conversation, recipes,
               model, input_tokens, output_tokens, cost, created_at
        FROM meal_plans ORDER BY id
    """)
    rows = [dict(r) for r in cur.fetchall()]
    for row in rows:
        for k in ("prefs", "recipes", "conversation"):
            if isinstance(row[k], str):
                row[k] = json.loads(row[k])

    cur.execute("SELECT id FROM recipes")
    existing_ids = {r["id"] for r in cur.fetchall()}

    cur.close(); conn.close()
    return rows, existing_ids

def generate_unique_id(used_ids, existing_db_ids):
    """Generate a random int that is NOT in the recipes table and NOT already assigned."""
    while True:
        new_id = random.randint(100_000, 9_999_999)
        if new_id not in existing_db_ids and new_id not in used_ids:
            used_ids.add(new_id)
            return new_id

def deduplicate_recipes(meal_plans):
    """
    Across all meal plans, collect unique recipes by name (case-insensitive).
    Returns list of (recipe_dict, parent_meal_plan_dict).
    """
    seen_names = set()
    unique = []
    for mp in meal_plans:
        for rec in mp.get("recipes", []):
            name_key = rec.get("name", "").strip().lower()
            if name_key and name_key not in seen_names:
                seen_names.add(name_key)
                unique.append((rec, mp))
    return unique

# ─────────────────────────────────────────────────────────────
# SHEET 1 — Overview
# ─────────────────────────────────────────────────────────────
def sheet_overview(wb, meal_plans):
    ws = wb.active
    ws.title = "Overview"
    ws.sheet_view.showGridLines = False

    total_recipes = sum(len(mp["recipes"]) for mp in meal_plans)

    ws.merge_cells("A1:F1")
    ws["A1"].value = f"Meal Plans Export  —  {len(meal_plans)} plans  |  {total_recipes} total recipes (before dedup)"
    ws["A1"].fill  = fill(C_DARK_BLUE)
    ws["A1"].font  = fnt(bold=True, color=C_WHITE, size=14)
    ws["A1"].alignment = aln("center")
    ws.row_dimensions[1].height = 32

    set_hdr(ws.cell(2, 1), "MP ID",         C_DARK_BLUE)
    set_hdr(ws.cell(2, 2), "User ID",        C_DARK_BLUE)
    set_hdr(ws.cell(2, 3), "Model",          C_DARK_BLUE)
    set_hdr(ws.cell(2, 4), "Cost",           C_DARK_BLUE)
    set_hdr(ws.cell(2, 5), "Recipe Count",   C_DARK_BLUE)
    set_hdr(ws.cell(2, 6), "Created At",     C_DARK_BLUE)

    for ri, mp in enumerate(meal_plans, 3):
        bg = C_PALE_BLUE if ri % 2 == 0 else C_WHITE
        set_cel(ws.cell(ri, 1), mp["id"],                                    bg, bold=True, fg=C_DARK_BLUE)
        set_cel(ws.cell(ri, 2), mp["user_id"],                               bg, fg="333333")
        set_cel(ws.cell(ri, 3), mp["model"],                                 bg, italic=True, fg="777777")
        set_cel(ws.cell(ri, 4), round(mp["cost"], 6) if mp["cost"] else 0,   bg, fg="555555")
        set_cel(ws.cell(ri, 5), len(mp["recipes"]),                          bg, fg="333333", center=True)
        set_cel(ws.cell(ri, 6), str(mp["created_at"]),                       bg, fg="555555")

    widths(ws, {"A": 12, "B": 30, "C": 24, "D": 14, "E": 14, "F": 28})

# ─────────────────────────────────────────────────────────────
# SHEET 2 — recipes table
# ─────────────────────────────────────────────────────────────
def sheet_recipes(wb, unique_recipes, recipe_id_map_A, recipe_id_map_B):
    ws = wb.create_sheet("recipes")
    ws.sheet_view.showGridLines = False

    cols = [
        ("generated_id",      False), ("recipe_name",        False),
        ("cuisines",          False), ("servings",            False),
        ("cooking_time",      False), ("cooking_complexity",  False),
        ("calories",          False), ("fat",                 False),
        ("carbs",             False), ("protein",             False),
        ("meal_type",         False), ("ingredients",         False),
        ("instructions",      False), ("diet_tags",           False),
        ("contains_tags",     False), ("source",              False),
        ("source_id",         False), ("created_at",          False),
        ("short_summary",      True), ("popularity_score",    True),
        ("image_url",          True), ("likes",               True),
        ("tags",               True), ("focus_categories",    True),
        ("updated_at",         True), ("search_vector",       True),
        ("ingredients_tsv",    True),
    ]
    total_cols = len(cols)

    ws.merge_cells(f"A1:{get_column_letter(total_cols)}1")
    ws["A1"].value = f"recipes  —  {len(unique_recipes)} unique rows (duplicates removed)"
    ws["A1"].fill  = fill(C_DARK_BLUE)
    ws["A1"].font  = fnt(bold=True, color=C_WHITE, size=13)
    ws["A1"].alignment = aln("center")
    ws.row_dimensions[1].height = 28

    for ci, (col, is_null) in enumerate(cols, 1):
        set_hdr(ws.cell(2, ci), col, bg="9E9E9E" if is_null else C_MID_BLUE)
    ws.row_dimensions[2].height = 28

    for ri, (rec, mp) in enumerate(unique_recipes, 3):
        p     = mp["prefs"] or {}
        n     = rec.get("nutrition") or {}
        ingr  = rec.get("ingredients") or []
        instr = rec.get("instructions") or []
        mins  = parse_mins(rec.get("time"))

        name_key = rec.get("name", "").strip().lower()
        gen_id_A = recipe_id_map_A[name_key]   # unique ID for recipes.generated_id
        gen_id_B = recipe_id_map_B[name_key]   # shared ID for recipes.source_id

        s_ingr = json.dumps([{
            "name":   x.get("item", ""),
            "amount": parse_amt(x.get("amount", ""))[0],
            "unit":   parse_amt(x.get("amount", ""))[1],
            "price":  x.get("price", 0)
        } for x in ingr])

        s_instr = json.dumps([
            {"step": i + 1, "text": s} for i, s in enumerate(instr)
        ])

        bg = C_PALE_BLUE if ri % 2 == 0 else C_WHITE
        values = [
            gen_id_A,                               # generated_id  — ID A (unique to this sheet)
            rec.get("name"),
            json.dumps([rec.get("cuisine")]),
            rec.get("servings"),
            mins,
            complexity(mins),
            n.get("calories"),
            str(n.get("fat")),
            str(n.get("carbs")),
            str(n.get("protein")),
            json.dumps([rec.get("meal")]),
            s_ingr,
            s_instr,
            json.dumps([p.get("dietary") or ""]),
            json.dumps(p.get("allergies") or []),
            "meal_plans",
            gen_id_B,                               # source_id     — ID B (shared with recipes_normalized)
            str(mp["created_at"]),
            "NULL", 0, "NULL", 0, "NULL", "NULL",
            str(mp["created_at"]), "NULL", "NULL",
        ]

        for ci, (val, (_, is_null)) in enumerate(zip(values, cols), 1):
            set_cel(ws.cell(ri, ci), val,
                    bg="F5F5F5" if is_null else bg,
                    bold=(ci == 1),
                    fg="AAAAAA" if is_null else ("1B5E20" if ci == 1 else "333333"),
                    italic=is_null)
        ws.row_dimensions[ri].height = 42

    col_w = {
        "A": 14, "B": 28, "C": 16, "D": 10, "E": 13, "F": 18,
        "G": 10, "H": 8,  "I": 8,  "J": 9,  "K": 16,
        "L": 38, "M": 38, "N": 16, "O": 20, "P": 13,
        "Q": 12, "R": 22,
    }
    for i in range(19, total_cols + 1):
        col_w[get_column_letter(i)] = 16
    widths(ws, col_w)

# ─────────────────────────────────────────────────────────────
# SHEET 3 — recipes_normalized table
# ─────────────────────────────────────────────────────────────
def sheet_recipes_normalized(wb, unique_recipes, recipe_id_map_B):
    ws = wb.create_sheet("recipes_normalized")
    ws.sheet_view.showGridLines = False

    mapped_cols = [
        ("generated_id",               False),
        ("title",                      False),
        ("new_title",                  False),
        ("servings",                   False),
        ("ready_in_minutes",           False),
        ("cuisines",                   False),
        ("dish_types",                 False),
        ("vegetarian",                 False),
        ("diets",                      False),
        ("new_instructions",           False),
        ("new_extended_ingredients",   False),
        ("original_id",                False),
        ("source_name",                False),
        ("created_at",                 False),
    ]
    null_cols = [
        ("image",                       True), ("image_type",                  True),
        ("source_url",                  True), ("vegan",                        True),
        ("gluten_free",                 True), ("dairy_free",                   True),
        ("very_healthy",                True), ("cheap",                        True),
        ("very_popular",                True), ("sustainable",                  True),
        ("low_fodmap",                  True), ("weight_watcher_smart_points",  True),
        ("health_score",                True), ("preparation_minutes",          True),
        ("cooking_minutes",             True), ("aggregate_likes",              True),
        ("price_per_serving",           True), ("summary",                      True),
        ("occasions",                   True), ("wine_pairing",                 True),
        ("instructions",                True), ("analyzed_instructions",        True),
        ("spoonacular_score",           True), ("extended_ingredients",         True),
        ("updated_extended_ingredients",True), ("new_tags",                     True),
    ]
    all_cols = mapped_cols + null_cols
    total_cols = len(all_cols)

    ws.merge_cells(f"A1:{get_column_letter(total_cols)}1")
    ws["A1"].value = f"recipes_normalized  —  {len(unique_recipes)} unique rows (duplicates removed)"
    ws["A1"].fill  = fill(C_PURPLE)
    ws["A1"].font  = fnt(bold=True, color=C_WHITE, size=13)
    ws["A1"].alignment = aln("center")
    ws.row_dimensions[1].height = 28

    for ci, (col, is_null) in enumerate(all_cols, 1):
        set_hdr(ws.cell(2, ci), col, bg="9E9E9E" if is_null else "6A1B9A")
    ws.row_dimensions[2].height = 28

    null_defaults = {
        "vegan": False, "gluten_free": False, "dairy_free": False,
        "very_healthy": False, "cheap": False, "very_popular": False,
        "sustainable": False, "low_fodmap": False,
        "aggregate_likes": 0, "price_per_serving": "DERIVED",
    }

    for ri, (rec, mp) in enumerate(unique_recipes, 3):
        p     = mp["prefs"] or {}
        ingr  = rec.get("ingredients") or []
        instr = rec.get("instructions") or []
        mins  = parse_mins(rec.get("time"))
        is_veg = (p.get("dietary") or "").lower() == "vegetarian"

        name_key = rec.get("name", "").strip().lower()
        gen_id_B = recipe_id_map_B[name_key]   # shared ID — same as recipes.source_id

        new_ext = json.dumps([{
            "id":        -1,
            "name":      x.get("item", ""),
            "unit":      parse_amt(x.get("amount", ""))[1],
            "amount":    _safe_float(parse_amt(x.get("amount", ""))[0]),
            "original":  f"{x.get('amount','')} {x.get('item','')}",
            "nameClean": x.get("item", "").lower()
        } for x in ingr])

        new_instr = json.dumps([
            {"step": i + 1, "text": s} for i, s in enumerate(instr)
        ])

        bg = C_LIGHT_PURP if ri % 2 == 0 else C_WHITE
        mapped_vals = [
            gen_id_B,                               # generated_id  — ID B (same as recipes.source_id)
            rec.get("name"),
            rec.get("name"),
            rec.get("servings"),
            mins,
            "{" + (rec.get("cuisine") or "") + "}",
            "{" + (rec.get("meal") or "").lower() + "}",
            is_veg,
            "{" + (p.get("dietary") or "") + "}",
            new_instr,
            new_ext,
            mp["id"],
            "meal_plans_ai",
            str(mp["created_at"]),
        ]
        null_vals = [null_defaults.get(col, "NULL") for col, _ in null_cols]
        all_vals  = mapped_vals + null_vals

        for ci, (val, (_, is_null)) in enumerate(zip(all_vals, all_cols), 1):
            set_cel(ws.cell(ri, ci), val,
                    bg="F5F5F5" if is_null else bg,
                    bold=(ci == 1),
                    fg="AAAAAA" if is_null else ("4A148C" if ci == 1 else "333333"),
                    italic=is_null)
        ws.row_dimensions[ri].height = 42

    col_w = {}
    wide_cols = {2, 3, 10, 11}
    for i, (col, is_null) in enumerate(all_cols, 1):
        letter = get_column_letter(i)
        col_w[letter] = 30 if i in wide_cols else (14 if is_null else 18)
    col_w["A"] = 14
    widths(ws, col_w)

# ─────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────
def main():
    print("Connecting to database...")
    meal_plans, existing_db_ids = fetch_meal_plans()
    print(f"Fetched {len(meal_plans)} meal plans  |  {len(existing_db_ids)} existing recipe IDs in DB")

    unique_recipes = deduplicate_recipes(meal_plans)
    total_raw = sum(len(mp["recipes"]) for mp in meal_plans)
    dupes_removed = total_raw - len(unique_recipes)
    print(f"Total recipes across all plans : {total_raw}")
    print(f"Unique recipes (after dedup)   : {len(unique_recipes)}")
    print(f"Duplicates removed             : {dupes_removed}")

    # Two separate ID maps per unique recipe:
    #   recipe_id_map_A → used only for recipes.generated_id
    #   recipe_id_map_B → used for recipes.source_id AND recipes_normalized.generated_id
    used_ids = set()
    recipe_id_map_A = {}
    recipe_id_map_B = {}
    for rec, _ in unique_recipes:
        name_key = rec.get("name", "").strip().lower()
        recipe_id_map_A[name_key] = generate_unique_id(used_ids, existing_db_ids)
        recipe_id_map_B[name_key] = generate_unique_id(used_ids, existing_db_ids)

    wb = Workbook()

    print("Building Sheet 1: Overview...")
    sheet_overview(wb, meal_plans)

    print("Building Sheet 2: recipes table...")
    sheet_recipes(wb, unique_recipes, recipe_id_map_A, recipe_id_map_B)

    print("Building Sheet 3: recipes_normalized table...")
    sheet_recipes_normalized(wb, unique_recipes, recipe_id_map_B)

    out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "meal_plans_all_2.xlsx")
    wb.save(out)
    print(f"\n✅  Saved  →  {out}")

if __name__ == "__main__":
    main()